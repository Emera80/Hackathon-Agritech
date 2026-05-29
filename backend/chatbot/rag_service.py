import os
import csv
import json
import re
import unicodedata
import pickle
import time
from openpyxl import load_workbook
from pypdf import PdfReader

# Liste de synonymes agricoles pour élargir la recherche
SYNONYMS = {
    "olivier": ["olives", "zaitoun", "olea", "huilerie", "زيتون", "الزيتون", "sfax", "huile d'olive"],
    "blé": ["céréales", "قمح", "wheat", "farine", "semis", "الحبوب", "القمح", "orge", "triticale"],
    "eau": ["irrigation", "stress hydrique", "barrage", "puits", "sondage", "مياه", "الري", "pluviométrie", "pluie", "secadenord", "sonede"],
    "maladie": ["parasite", "insecte", "champignon", "traitement", "pesticide", "امراض", "حشرات", "phytosanitaire"],
    "sol": ["terre", "fertile", "engrais", "labourage", "analyse", "تربة", "سماد", "Nis", "NPK", "pédologique", "profil pédologique", "type de sol", "argileux", "sablonneux", "calcaire", "sablonneux", "limoneux", "sableux", "drainage", "texture du sol"],
    "agrumes": ["قوارص", "citron", "orange", "mandarine", "cap bon", "nabeul", "manzel bouzalfa", "soliman"],
    "investissement": ["subvention", "prime", "apia", "crédit", "financement", "projet", "smsa", "subvention agricole", "aide état"],
    "prix": ["marché", "cotation", "tarif", "valeur", "fruit", "légume", "prix de vente", "prix gros"],
    "climat": ["micro-climat", "altitude", "exposition au soleil", "ensoleillement", "microclimat", "froid", "chaleur", "humidité", "vent", "gel", "sirocco", "chili", "température"],
    "region": ["tunisie", "favorable", "défavorable", "zone à risque", "nord", "sud", "centre", "sahel", "kroumirie", "jerid", "mornag", "sidi bouzid", "kairouan", "jendouba", "beja", "kef", "siliana", "zaghouan", "nabeul", "bizerte", "monastir", "mahdia", "sfax", "gabes", "medenine", "tataouine", "tozeur", "kebili", "gouvernorat", "délégation"],
    "culture": ["plante", "semis", "récolte", "rendement", "variété", "culture maraîchère", "arboriculture", "viticulture"],
}

STOPWORDS = {"le", "la", "les", "un", "une", "des", "du", "de", "ce", "cette", "ces", "mon", "ma", "mes", "ton", "ta", "tes", "son", "sa", "ses", "notre", "votre", "leur", "pour", "dans", "sur", "avec", "est", "sont", "était", "étaient", "aux", "par"}

def normalize_text(text):
    """Supprime les accents et met en minuscule."""
    text = "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )
    return text.lower()

class AgriRAG:
    def __init__(self, data_dir='data', cache_file='rag_cache.pkl'):
        self.data_dir = data_dir
        self.cache_file = cache_file
        self.documents = []
        self.inverted_index = {} # Mot -> Liste d'indices de documents

    def load_data(self):
        if not os.path.exists(self.data_dir):
            alt_path = os.path.join('backend', self.data_dir)
            if os.path.exists(alt_path):
                self.data_dir = alt_path
            else:
                os.makedirs(self.data_dir, exist_ok=True)
                return
        
        # Tentative de chargement depuis le cache
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'rb') as f:
                    cache_data = pickle.load(f)
                    self.documents = cache_data.get('documents', [])
                    self.inverted_index = cache_data.get('inverted_index', {})
                    if self.documents:
                        print(f"RAG: {len(self.documents)} documents chargés depuis le cache.")
                        return
            except Exception as e:
                print(f"RAG: Erreur chargement cache: {e}")

        print("RAG: Initialisation de la base de données (cela peut prendre du temps)...")
        start_time = time.time()
        
        # Convertir en chemin absolu pour éviter les problèmes de chemins relatifs longs
        abs_data_dir = os.path.abspath(self.data_dir)
        
        # Sur Windows, utiliser le préfixe \\?\ pour supporter les chemins de plus de 260 caractères
        if os.name == 'nt' and not abs_data_dir.startswith('\\\\?\\'):
            abs_data_dir = '\\\\?\\' + abs_data_dir

        for root, dirs, files in os.walk(abs_data_dir):
            for filename in files:
                if filename.startswith('._'): continue
                path = os.path.join(root, filename)
                
                # On garde le nom de fichier court pour la source, mais le path complet pour le chargement
                if filename.endswith('.csv'):
                    self._load_csv(path, filename)
                elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                    self._load_xlsx(path, filename)
                elif filename.endswith('.pdf'):
                    self._load_pdf(path, filename)
        
        self._build_inverted_index()
        self._save_cache()
        
        end_time = time.time()
        print(f"RAG: {len(self.documents)} documents indexés en {end_time - start_time:.2f}s.")

    def _build_inverted_index(self):
        """Construit un index inversé simple pour accélérer la recherche."""
        self.inverted_index = {}
        for idx, doc in enumerate(self.documents):
            # On indexe les mots du contenu normalisé et du titre
            text = doc['norm_content'] + " " + normalize_text(doc['source'])
            words = set(re.findall(r'\w+', text))
            for word in words:
                if len(word) > 2 and word not in STOPWORDS:
                    if word not in self.inverted_index:
                        self.inverted_index[word] = []
                    self.inverted_index[word].append(idx)

    def _save_cache(self):
        try:
            with open(self.cache_file, 'wb') as f:
                pickle.dump({
                    'documents': self.documents,
                    'inverted_index': self.inverted_index
                }, f)
        except Exception as e:
            print(f"RAG: Erreur sauvegarde cache: {e}")

    def _load_csv(self, path, filename):
        try:
            with open(path, mode='r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                chunk_size = 50
                current_chunk = []
                
                for i, row in enumerate(reader):
                    current_chunk.append(", ".join(row))
                    if len(current_chunk) >= chunk_size:
                        content = f"Fichier: {filename}\nColonnes: {header}\nDonnées (Lignes {i-chunk_size+1} à {i}):\n" + "\n".join(current_chunk)
                        self.documents.append({"source": filename, "content": content, "norm_content": normalize_text(content)})
                        current_chunk = []
                    if i > 2000: break # Augmentation de la limite
                
                if current_chunk:
                    content = f"Fichier: {filename}\nColonnes: {header}\nDonnées (Dernières lignes):\n" + "\n".join(current_chunk)
                    self.documents.append({"source": filename, "content": content, "norm_content": normalize_text(content)})
        except Exception as e:
            print(f"Error loading CSV {filename}: {e}")

    def _load_xlsx(self, path, filename):
        try:
            wb = load_workbook(path, data_only=True)
            for sheet in wb.worksheets:
                rows = list(sheet.iter_rows(max_row=1000, values_only=True))
                header = rows[0] if rows else []
                data_rows = rows[1:]
                
                chunk_size = 40
                for i in range(0, len(data_rows), chunk_size):
                    chunk = data_rows[i:i+chunk_size]
                    text = [f"Fichier: {filename}, Feuille: {sheet.title}, Colonnes: {header}"]
                    for r in chunk:
                        if any(r):
                            text.append(", ".join([str(c) if c is not None else "" for c in r]))
                    content = "\n".join(text)
                    self.documents.append({"source": filename, "content": content, "norm_content": normalize_text(content)})
        except Exception as e:
            print(f"Error loading XLSX {filename}: {e}")

    def _load_pdf(self, path, filename):
        try:
            reader = PdfReader(path)
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text.strip():
                    content = f"Fichier PDF: {filename}\n--- Page {i+1} ---\n{page_text}"
                    self.documents.append({"source": filename, "content": content, "norm_content": normalize_text(content)})
                if i > 50: break # Augmentation de la limite de pages
        except Exception as e:
            print(f"Error loading PDF {filename}: {e}")

    def search(self, query, top_k=3): # Réduction du top_k par défaut pour éviter 429
        norm_query = normalize_text(query)
        words = [w for w in re.findall(r'\w+', norm_query) if len(w) > 2 and w not in STOPWORDS]
        
        if not words:
            return []
            
        # Extension de la requête avec les synonymes (eux aussi normalisés)
        extended_words = set(words)
        for w in words:
            if w in SYNONYMS:
                for syn in SYNONYMS[w]:
                    extended_words.add(normalize_text(syn))
        
        # Filtrage des documents via l'index inversé
        candidate_indices = set()
        for word in extended_words:
            if word in self.inverted_index:
                candidate_indices.update(self.inverted_index[word])
        
        if not candidate_indices:
            return []

        scored_docs = []
        for idx in candidate_indices:
            doc = self.documents[idx]
            score = 0
            content = doc['norm_content']
            filename = normalize_text(doc['source'])
            
            for word in extended_words:
                # Occurrence dans le contenu (plus de poids pour les termes originaux que les synonymes)
                weight = 10 if word in words else 3
                
                count = content.count(word)
                score += count * weight
                
                # Gros bonus pour présence dans le titre du fichier
                if word in filename:
                    score += 100 
                
                # Bonus pour correspondance exacte (mot entier)
                if re.search(rf'\b{word}\b', content):
                    score += 20
            
            if score > 0:
                scored_docs.append((score, doc))
        
        scored_docs.sort(key=lambda x: x[0], reverse=True)
        
        results = []
        for score, doc in scored_docs[:top_k]:
            # Tronquer le contenu s'il est trop massif
            content = doc['content']
            if len(content) > 3000:
                content = content[:3000] + "... [Tronqué pour économiser des ressources]"
                
            results.append(f"--- Source: {doc['source']} (Score: {score}) ---\n{content}")
            
        return results

    def get_ready_context(self, query, has_media=False):
        """Prépare un contexte RAG formaté pour l'IA, optimisé selon la présence de médias."""
        top_k = 2 if has_media else 4
        results = self.search(query, top_k=top_k)
        
        if not results:
            return None
            
        return "\n\n".join(results)

# Singleton pour éviter de recharger à chaque fois si possible
_rag_instance = None

def get_rag():
    global _rag_instance
    if _rag_instance is None:
        _rag_instance = AgriRAG()
        _rag_instance.load_data()
    return _rag_instance
